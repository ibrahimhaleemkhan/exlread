#EXl-read
#read from Excel files from the B Row and plots a simple graph in seconds

#made By Ibrahim Haleem Khan
#github.com/ibrahimhaleemkhan
#more support coming soon
#more graphs Coming soon
#more rows coming soon

#Used MATPLOTLIB ,OPENPYXL and NUMPY (Under their Terms) 

try :
	import openpyxl, time
	import numpy as np
	import matplotlib.pyplot as plt
except Importerror:
	print "Please Installed Required Modules"
	print "1. Openpyxl"
	print "2. numpy"
	print "3. matplotlib"


def main():
	while True:
		graphside2=[]
		graphside1=[]
		print " \t \t \t WELCOME \n"
		print "\t \t \t------- | TO EXL-READ |-------\n \n \n"

		print "\tONLY ROW B WITH NUMERIC VALUES ARE SUPPORTED like B1, B2 , B3 \n "
		print "\t \t \tMORE SUPPORT COMING SOON! \n "

		print "\n"
		filename=raw_input("Enter The FIlLENAME(in.xlsx)Format: \n ")
		try:
			wb = openpyxl.load_workbook(filename)
		except:
			print "\t The file Does not exists, loading Default file \n "
			wb = openpyxl.load_workbook('random.xlsx')
			pass


		#current working sheet
		
		try:
			sheetselect = raw_input("\t \t \t Please Enter the Sheetname \n \t  You want  Graphical Data From (AS Sheet1 , Sheet2 or Sheet3 etc: \n")
			sheet = wb.get_sheet_by_name(sheetselect)
		except:
			print "\t \t \t Invalid sheet using Sheet1 as a main source \n \n "
			sheet = wb.get_sheet_by_name('Sheet1')

		
			
				
			

		print "\t \t \t Reading Current Sheet as Sheet \n"
		maxrow= sheet.max_row
		maxcol=sheet.max_column

		print "\t \t \t Max No Of Rows are :" , maxrow
		print "\t \t \t Max No of Cols are :" , maxcol
		print "\t \t \t _________________________________ \n "


		for i in range(maxrow):
			graphside2.append(i)
			

		for j in range(maxrow):
			p='B'+ str(j)
			if (j==0):
				j=1
			else:
				pass
			p=sheet.cell(row=j,column=2).value
			graphside1.append(p)
			


		m=np.array(graphside1)
		n=np.array(graphside2)

		#for plotting
		plt.plot(n,m)
		xname =raw_input("\t \t \tPlease Enter a label for X axis \n")
		plt.xlabel(xname)
		yname = raw_input("\t \t \tPlease Enter a label for y axis \n ")
		plt.ylabel(yname)
		heading = raw_input("\t \t \tPlease Give the graph a proper Heading \n ")
		plt.title(heading)
		plt.grid(True)
		print "\t Your File is saved as graph.png in the current directory \n"
		plt.savefig("graph.png")
		time.sleep(1)
		plt.show()
		

main()

